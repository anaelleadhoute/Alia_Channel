"""
Alia Video Pipeline
-------------------
Takes a video idea from the Excel by rank number and produces a talking video:
1. Claude generates a French script
2. ElevenLabs converts it to audio (French)
3. Flux generates a scene image of Alia in a cozy studio
4. Kling animates multiple 10-sec clips from that scene
5. ffmpeg stitches clips + merges French audio
6. Final video saved to output/ folder

Usage:
    python scripts/generate_video.py --rank 1
"""

import argparse
import os
import math
import subprocess
import tempfile
import requests
import anthropic
import fal_client
import openpyxl
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
ELEVENLABS_API_KEY = os.getenv("ELEVENLABS_API_KEY")
FAL_KEY = os.getenv("FAL_KEY")
os.environ["FAL_KEY"] = FAL_KEY or ""

EXCEL_PATH = Path(__file__).parent.parent / "Alia_100_Evergreen_Video_Ideas.xlsx"
AVATAR_PATH = Path(__file__).parent.parent / "images" / "alia_human_avatar.png"
OUTPUT_DIR = Path(__file__).parent.parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

ELEVENLABS_VOICE_ID = "cgSgspJ2msm6clMCkdW9"  # Jessica - Playful, Bright, Warm
CLIP_DURATION = 10  # seconds per Kling clip


def load_video_idea(rank: int) -> dict:
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == rank:
            return dict(zip(headers, row))
    raise ValueError(f"No video idea with rank {rank}")


def generate_script(idea: dict) -> str:
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    prompt = f"""Tu es Alia, une assistante virtuelle francophone qui aide les olim en Israël.
Tu parles directement à la caméra dans une vidéo Instagram de 30 à 45 secondes.

Voici le sujet :
- Titre : {idea['Title']}
- Accroche : {idea['Hook']}
- Problème : {idea['User Problem']}
- Bénéfice : {idea['Benefit']}
- Résumé : {idea['Summary']}

Rédige un script parlé en français, naturel et chaleureux, de 30 à 45 secondes (80-120 mots).
Structure :
1. Accroche percutante (5 sec)
2. Le problème que tu connais peut-être (10 sec)
3. La solution concrète (15 sec)
4. Call to action : rejoins la communauté Alia (10 sec)

Réponds uniquement avec le texte du script, sans titre ni indication de mise en scène."""

    response = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=400,
        messages=[{"role": "user", "content": prompt}],
    )
    return response.content[0].text.strip()


def generate_audio(script: str, output_path: Path) -> float:
    url = f"https://api.elevenlabs.io/v1/text-to-speech/{ELEVENLABS_VOICE_ID}"
    headers = {
        "xi-api-key": ELEVENLABS_API_KEY,
        "Content-Type": "application/json",
    }
    payload = {
        "text": script,
        "model_id": "eleven_multilingual_v2",
        "voice_settings": {
            "stability": 0.5,
            "similarity_boost": 0.75,
            "style": 0.3,
            "use_speaker_boost": True,
        },
    }
    response = requests.post(url, headers=headers, json=payload)
    response.raise_for_status()
    output_path.write_bytes(response.content)

    # Get audio duration via ffprobe
    result = subprocess.run(
        ["ffprobe", "-v", "error", "-show_entries", "format=duration",
         "-of", "default=noprint_wrappers=1:nokey=1", str(output_path)],
        capture_output=True, text=True
    )
    duration = float(result.stdout.strip())
    print(f"  Audio saved: {output_path} ({duration:.1f}s)")
    return duration


def upload_bytes_to_fal(data: bytes, content_type: str) -> str:
    return fal_client.upload(data, content_type=content_type)


def generate_scene_image(output_path: Path) -> str:
    print("  Generating cozy studio scene with Alia (Flux)...")
    avatar_data = AVATAR_PATH.read_bytes()
    avatar_url = upload_bytes_to_fal(avatar_data, "image/png")

    result = fal_client.run(
        "fal-ai/flux/dev/image-to-image",
        arguments={
            "image_url": avatar_url,
            "prompt": """The same woman, sitting in a cozy home studio.
She is positioned in front of a professional podcast microphone on a desk.
Warm soft lighting from the side, bookshelves with plants and books in the background.
Modern cozy aesthetic, warm amber tones. She is looking directly at the camera with a warm smile.
Photorealistic, high quality, vertical portrait orientation.""",
            "strength": 0.55,
            "num_inference_steps": 28,
            "guidance_scale": 3.5,
        },
    )
    image_url = result["images"][0]["url"]
    image_data = requests.get(image_url).content
    output_path.write_bytes(image_data)
    print(f"  Scene image saved: {output_path}")
    return image_url


def generate_kling_clip(scene_image_url: str, clip_index: int, output_path: Path) -> Path:
    print(f"  Generating clip {clip_index + 1} with Kling...")

    result = fal_client.run(
        "fal-ai/kling-video/v2.6/pro/image-to-video",
        arguments={
            "image_url": scene_image_url,
            "prompt": """The woman is speaking naturally to camera, animated and expressive.
Natural lip movement, head nodding slightly, blinking naturally.
Cozy studio background stays consistent. Warm lighting. Influencer style.""",
            "negative_prompt": "transition, fade, cut, zoom, blurry, distorted face, static frozen",
            "duration": "10",
            "aspect_ratio": "9:16",
            "cfg_scale": 0.5,
        },
    )

    video_url = result["video"]["url"]
    video_data = requests.get(video_url).content
    output_path.write_bytes(video_data)
    print(f"  Clip {clip_index + 1} saved: {output_path}")
    return output_path


def stitch_clips_with_audio(clip_paths: list[Path], audio_path: Path, output_path: Path):
    print("  Stitching clips and merging audio with ffmpeg...")

    with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", delete=False) as f:
        for clip in clip_paths:
            f.write(f"file '{clip.absolute()}'\n")
        concat_file = f.name

    stitched = OUTPUT_DIR / "stitched_temp.mp4"
    subprocess.run([
        "ffmpeg", "-y", "-f", "concat", "-safe", "0",
        "-i", concat_file,
        "-c", "copy", str(stitched)
    ], check=True, capture_output=True)

    # Merge audio — trim video to audio length
    subprocess.run([
        "ffmpeg", "-y",
        "-i", str(stitched),
        "-i", str(audio_path),
        "-map", "0:v:0",
        "-map", "1:a:0",
        "-c:v", "copy",
        "-c:a", "aac",
        "-shortest",
        str(output_path)
    ], check=True, capture_output=True)

    stitched.unlink(missing_ok=True)
    os.unlink(concat_file)
    print(f"  Final video saved: {output_path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--rank", type=int, required=True)
    args = parser.parse_args()

    print(f"\n=== Alia Video Pipeline — Rank #{args.rank} ===\n")

    # 1. Load idea
    print("1. Loading video idea...")
    idea = load_video_idea(args.rank)
    print(f"   Title: {idea['Title']}")

    slug = idea['Title'][:40].replace(" ", "_").replace("/", "-").lower()
    audio_path = OUTPUT_DIR / f"rank_{args.rank}_{slug}.mp3"
    scene_path = OUTPUT_DIR / f"rank_{args.rank}_{slug}_scene.png"
    video_path = OUTPUT_DIR / f"rank_{args.rank}_{slug}.mp4"

    # 2. Generate script
    print("\n2. Generating French script with Claude...")
    script = generate_script(idea)
    print(f"\n--- SCRIPT ---\n{script}\n--------------\n")

    # 3. Generate audio + get duration
    print("3. Generating French audio with ElevenLabs...")
    audio_duration = generate_audio(script, audio_path)

    # 4. Generate scene image with Alia in cozy studio
    print("\n4. Generating cozy studio scene image...")
    scene_image_url = generate_scene_image(scene_path)

    # 5. Generate enough Kling clips to cover audio duration
    num_clips = math.ceil(audio_duration / CLIP_DURATION)
    print(f"\n5. Generating {num_clips} Kling clips ({audio_duration:.1f}s audio)...")
    clip_paths = []
    for i in range(num_clips):
        clip_path = OUTPUT_DIR / f"rank_{args.rank}_{slug}_clip_{i}.mp4"
        generate_kling_clip(scene_image_url, i, clip_path)
        clip_paths.append(clip_path)

    # 6. Stitch clips + merge audio
    print("\n6. Stitching clips and merging French audio...")
    stitch_clips_with_audio(clip_paths, audio_path, video_path)

    # Cleanup clips
    for clip in clip_paths:
        clip.unlink(missing_ok=True)

    print(f"\n✅ Done! Video saved to:\n   {video_path}\n")


if __name__ == "__main__":
    main()
